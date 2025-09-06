#from openpyxl import load_workbook
import os
import openpyxl as op
import simplekml
import csv
from proj4 import JTSK_to_WGS
from _utils import dirEntries
from _settings import *


#PATHBASEINDB = r'f:\aaa\PROGRAM\python\vrty2\Vrty_Geo5'
#WEBTOPDIR = r'http://172.16.0.2/dokumenty/Vrty_Geo5'
KMLNAME = TOPDIR + r'\Geo5N.kml'
_KML = '' #global variable initialized in main

def get_cells_dict(sheet):
    '''fn dostane sheet z ktorého vráti zoznam dictionaries
    s dátami. Je to pokus, v tejto etape budeme načítavať
    len "FieldTests"'''
    rows =[]
    i = True
    for row in sheet.rows:
        row = [cell.value for cell in row]
        if i:
            header = row
            i = False
        else: 
            rows.append(dict(zip(header, row)))
            #rows.append(zip(header, row))
    # for row in rows:
    #     print(row)
    return rows

def get_URL_uloha(vrtname, wbname):
    (head, tail) = os.path.split(wbname)
    #print("head-tail ",head, tail)
    pdfname = head+'\\'+vrtname+'.pdf'
    # if os.path.isfile(pdfname):
    #     print(pdfname, 'found')
    # else:
    #     print(pdfname, 'not found')
    urlname = pdfname.replace(PATHBASEINDB, WEBTOPDIR).replace('\\', '/')
    uloha = head.replace(PATHBASEINDB+'\\', '')
    #print(urlname)
    return(urlname, uloha)

def kmlwrite_one_point_balloon(vrt):
    '''Používa global _KML'''
    global _KML
    btext = 'čís:{}<br>'.format(vrt['Názov skúšky'])+\
            'úloha:{}<br>'.format(vrt['Úloha'])+\
            'výška:{}<br>'.format(vrt['Súradnica Z'])+\
            'dokumentoval:{}<br>'.format(vrt['Dokumentoval'])+\
            '<a href="{}"> PDF </a>'.format(vrt['URL'])

    pt = _KML.newpoint(name=vrt['Názov skúšky'], coords=[(vrt['Lon'],vrt['Lat'])])
    #print(vrt['Názov skúšky'], vrt['Súradnica X'], vrt['Súradnica Y'],str(vrt['Lat']),str(vrt['Lon']), vrt['URL'])
    pt.description = vrt['Názov skúšky']
    pt.balloonstyle.text = btext
    pt.style.iconstyle.color ='ff00ff00' # aabbggrr

def kmlwrite_one_point_extended(vrt):
    '''Používa global _KML'''
    global _KML
    
    pt = _KML.newpoint(name=vrt['Názov skúšky'], coords=[(vrt['Lon'],vrt['Lat'])])
    pt.extendeddata.newdata(name='birds', value=400, displayname="Bird Species")
    pt.extendeddata.newdata(name='aviaries', value=100, displayname="Aviaries")
    pt.extendeddata.newdata(name='visitors', value=10000, displayname="Annual Visitors")
    

def process_workbook(wbname):
    wb = op.load_workbook(wbname)
    sheet = wb['FieldTests']
    if sheet:
        vrty = get_cells_dict(sheet)
    else:
        print(wbname, "neni")
        return
    retval = []
    for vrt in vrty:
        [vrt['Lat'], vrt['Lon']] =JTSK_to_WGS(str(vrt['Súradnica X']),str(vrt['Súradnica Y'])) #pridáme WGS
        (vrt['URL'], vrt['Úloha']) = get_URL_uloha(vrt['Názov skúšky'], wbname)
        try:
            #kmlwrite_one_point_extended(vrt)
            kmlwrite_one_point_balloon(vrt)
            retval.append(vrt)
        except Exception as err:
            print("READER Exception:", dir, f"{err=}, {type(err)=}") 
    return list(retval)
    
def write_csv(wbname):
    wb = op.load_workbook(wbname)
    sheet = wb['FieldTests']
    if sheet:
        vrty = get_cells_dict(sheet)
    else:
        print(wbname, "neni")
        return
    
    for vrt in vrty:
        [vrt['Lat'], vrt['Lon']] =JTSK_to_WGS(str(vrt['Súradnica X']),str(vrt['Súradnica Y'])) #pridáme WGS
        (vrt['URL'], vrt['Úloha']) = get_URL_uloha(vrt['Názov skúšky'], wbname)
    return vrty    
   
HEADER1 = ['Názov skúšky',	'Template',	'Súradnica X', 'Súradnica Y', 'Súradnica Z','Posun počiatku',\
        'Súradnica Z pažnice', 'Príloha č.', 'Vrtmajster', 'Dokumentoval', 'Dátum zač.', 'Dátum kon.']
HEADER2 = ['Názov skúšky',	'Template',	'Súradnica X', 'Súradnica Y', 'Súradnica Z','Posun počiatku',\
        'Príloha č.', 'Vrtmajster',	'Dokumentoval',	'Dátum zač.', 'Dátum kon.',\
        'Poznámky k vrtu	Dáta - Skúška|Vrtná súprava']
# HEADER ='''Vrt;File;Uloha;Priloha;Ucel;Firma;Lokalita;Okres;Kraj;JTSKX;\
# JTSKY;Hteren;Hpaznica;Dielo;Etapa;Obstaravatel;Vrtal;Suprava;Vrtmajster;Doba;\
# Geolog;Mierka;Hlbka;void;Lat;Lon;HPV;Na;PDF\n'''

HEADER ='''Vrt;Uloha;JTSKX;JTSKY;Hteren;Vrtal;Geolog;Hlbka;Lat;Lon;PDF\n'''
SEP = ';'

wblist = dirEntries(PATHBASEINDB,True, 'xls', 'xlsx')
_KML = simplekml.Kml()
# print(wblist)    
vrtyDict = list()
for xlsx in wblist:
    vrty = (process_workbook(xlsx))
    # print(vrty) 
    for dic in vrty:
        vrtyDict.append(dic)
        
    # for dic in vrtyDict:
    #     print('\n'+ str(dic))
    # #print(vrtyDict)

    # f= open('GEO5.csv', 'w')  
    # f.write(';'.join(vrtyDict[0].keys())+'\n')
    # for vrt in vrtyDict:
    #     vals = map(str,vrt.values())
    #     vals = [x.replace('\n', ' ') for x in vals]
    #     f.write(';'.join(vals)+'\n')
    f= open(GEO5QGIS, 'w')  
    f.write(HEADER)
    for vrt in vrtyDict:
         vals = map(str,vrt.values())
         vals = [x.replace('\n', ' ') for x in vals]
         f.write('{};{};{};{};{};{};{};{};{};{};{}\n'.format(vrt['Názov skúšky'], vrt['Úloha'], vrt['Súradnica X'], vrt['Súradnica Y'], vrt['Súradnica Z']\
                 , vrt['Vrtmajster'], vrt['Dokumentoval'], vrt['Úloha'], vrt['Lat'], vrt['Lon'], vrt['URL'] ))
       


_KML.save(KMLNAME)
#print(vrtyDict)


print('DONE')


 