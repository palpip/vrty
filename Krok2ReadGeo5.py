#from openpyxl import load_workbook
import os
import openpyxl as op
import simplekml
import csv
import logging
import pandas as pd
from proj4 import JTSK_to_WGS
from _utils import dirEntries
from _funcs import chkdirs
from _settings import *
import warnings

chkdirs()

logger=logging.getLogger('vrt')
logger.addHandler(logging.FileHandler(KROK2LOGFILE, mode='a'))
logger.addHandler(logging.StreamHandler())
logger_pdf = logging.getLogger('pdf')
logger_pdf.addHandler(logging.StreamHandler())
logger_pdf.addHandler(logging.FileHandler(KROK2PDF, mode='a'))


# KMLNAME = TOPDIR + r'\Geo5N.kml'
# _KML = '' #global variable initialized in main

def get_cells_dict(sheet):
    '''fn dostane sheet z ktorého vráti zoznam dictionaries
    s dátami. Je to pokus, v tejto etape budeme načítavať
    len "FieldTests"'''
    rows =[]
    i = True
    j = 0
    header = []
    for row in sheet.iter_rows():
        row = [cell.value for cell in row]
        if i:
            header = row
            i = False
        else: 
            rows.append(dict(zip(header, row)))
            j += 1
            #rows.append(zip(header, row))
    # for row in rows:
    #      print(row)
    # print(j)
    return rows

def get_URL_uloha(vrtname, wbname):
    (head, tail) = os.path.split(wbname)
    #print("head-tail ",head, tail)
    pdfname = head+'\\'+vrtname+'.pdf'
    if os.path.isfile(pdfname):
        # print(pdfname, 'found')
        pass
    else:
        logger_pdf.warning (vrtname + ' : ' + pdfname + " nenajdene, polozka pripravena")

    urlname = pdfname.replace(PATHBASEINDB, WEBTOPDIR).replace('\\', '/')
    uloha = head.replace(PATHBASEINDB+'\\', '')
    return(urlname, uloha)


def get_hlbky_dict(sheet):
    '''vytvori dictionary vrt:maxhlbka a vrati ho'''
    res = {}
    for row in sheet.rows:
        hlbka = row[2].value
        vrt = row[0].value
        # print(vrt,hlbka)
        if vrt not in res:
            res[vrt] = hlbka
        elif  vrt != None and res[vrt] < hlbka:    # iterator row ide niekedy az do prazdneho riadku. To je chyba
            res[vrt] = hlbka
    return res

def process_workbook(wbname):
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = op.load_workbook(wbname)

    print(wbname)
    if 'FieldTests' not in wb.sheetnames:
        logger.error (f'{wbname} nie je z Geo5')
        return []

    sheet = wb['FieldTests']
    if sheet:
        vrty = get_cells_dict(sheet)
    else:
        print(wbname, "neni")
        return
    hlbky = get_hlbky_dict(wb['Dáta - Skúška|Vŕtanie'])
    retval = []
    for vrt in vrty:
        #eliminuj chyby ak načíta prázdny riadok
        if vrt['Názov skúšky']:
            # print(wbname, vrt['Názov skúšky'])
            if vrt['Súradnica X'] and vrt['Súradnica Y']:
                if vrt['Súradnica X'] < vrt['Súradnica Y']:
                    vrt['Súradnica X'] , vrt['Súradnica Y'] = vrt['Súradnica Y'] , vrt['Súradnica X']
                    pass
                [vrt['Lat'], vrt['Lon']] =JTSK_to_WGS(str(vrt['Súradnica X']),str(vrt['Súradnica Y']))
            else:
                vrt['Lat'] = vrt['Lon'] = 0 #pridáme WGS
                logger.error('suradnice v %s vo vrte %s' % (wbname, vrt['Názov skúšky']))
            (vrt['URL'], vrt['Úloha']) = get_URL_uloha(vrt['Názov skúšky'], wbname)
            if vrt['Názov skúšky'] in hlbky:
                vrt['Hĺbka'] = hlbky[vrt['Názov skúšky']]
            else:
                vrt['Hĺbka'] = 0
            # print(vrt)
            try:
                # kmlwrite_one_point_balloon(vrt)
                retval.append(vrt)
            except Exception as err:
                # print("READER Exception:", wbname, f"{err=}, {type(err)=}") 
                logger.error("Exception: %s %s %s", wbname, err, type(err))

    return list(retval)
    
def write_csv(wbname):
    wb = op.load_workbook(wbname)
    sheet = wb['FieldTests']
    if sheet:
        vrty = get_cells_dict(sheet)
    else:
        print(wbname, "neni")
        return
    
    for vrt in vrty:
        if vrt['Súradnica X'] < vrt['Súradnica Y']:
            vrt['Súradnica X'] , vrt['Súradnica Y'] = vrt['Súradnica Y'] , vrt['Súradnica X']
            pass
        [vrt['Lat'], vrt['Lon']] =JTSK_to_WGS(str(vrt['Súradnica X']),str(vrt['Súradnica Y'])) #pridáme WGS
        (vrt['URL'], vrt['Úloha']) = get_URL_uloha(vrt['Názov skúšky'], wbname)
    return vrty    
   

HEADER ='''Vrt;Uloha;JTSKX;JTSKY;Hteren;Vrtal;Geolog;Hlbka;Lat;Lon;URL;\n'''

#len xlsx, openpyxl nepodporuje xls
wblist = dirEntries(TOPDIR,True,  'xlsx')
# print('wblist:', wblist)
# _KML = simplekml.Kml()
vrtyDict = list()
vrty = list()

df = pd.DataFrame()    
f = open(GEO5QGIS, 'w')
f.write(HEADER)
    

for xlsx in wblist:
    
    if xlsx != EXCELWB: #nie vysledkovy excel, len excely v podadresaroch vrtov
        #print(xlsx) 
        vrty = (process_workbook(xlsx))
    for dic in vrty:
        if dic: vrtyDict.append(dic)
    for vrt in vrtyDict:
        try:
            vals = map(str,vrt.values())
            vals = [x.replace('\n', ' ') for x in vals]
            if not 'Vrtmajster' in vrt: vrt['Vrtmajster'] = '-'
            if not 'Dokumentoval' in vrt: vrt['Dokumentoval'] = '-'
            
            f.write('{};{};{};{};{};{};{};{};{};{};{}\n'.format(vrt['Názov skúšky'], vrt['Úloha'], vrt['Súradnica X'], vrt['Súradnica Y'], vrt['Súradnica Z']\
                    , vrt['Vrtmajster'], vrt['Dokumentoval'], vrt['Hĺbka'], vrt['Lat'], vrt['Lon'], vrt['URL'] ))
            #  print('{};{};{};{};{};{};{};{};{};{};{}\n'.format(vrt['Názov skúšky'], vrt['Úloha'], vrt['Súradnica X'], vrt['Súradnica Y'], vrt['Súradnica Z']\
            #          , vrt['Vrtmajster'], vrt['Dokumentoval'], vrt['Hĺbka'], vrt['Lat'], vrt['Lon'], vrt['URL'] ))
            vrt_the_dict = pd.DataFrame({'name':vrt.keys(), 'value':vrt.values()}).transpose()
            
        except Exception as err:
            logger.error("Exception: %s %s %s", vrt, err, type(err))
    try:
        # df1 = pd.read_excel(xlsx, sheet_name = 'FieldTests', engine='calamine')
        # df = pd.concat([df1, df])
        pass
    except Exception as err:
        logger.error("Exception sheet: %s %s %s", xlsx, err, type(err))

# _KML.save(KMLNAME)
#print(vrtyDict)

f.close
print('DONE')


 