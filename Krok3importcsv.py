import csv
import os
import logging
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from  _settings import *

# VRTCSVQGIS = 'vrtyGIS.csv'    
# VRTLOZCSVQGIS = 'vrtlozGIS.csv'
# GEO5QGIS = 'GEO5QGIS.csv'

# logging.basicConfig(filename = KROK3LOGFILE, level=logging.INFO, filemode='w')
# logger=logging.getLogger()

logger=logging.getLogger('vrt')
logger.addHandler(logging.FileHandler(KROK3LOGFILE, mode='a'))

TOPDIR = TOPDIR + '\\'
dest_filename = TOPDIR + r"vrty.xlsx"
wb = Workbook()
# wb.create_sheet(VRTCSVQGIS)
# wb.create_sheet(VRTLOZCSVQGIS)

def create_xls_from_cvs(fn, shname, shnum):
    print(fn)
    f = open(fn)
    csv.register_dialect('semicolons', delimiter=';')
    reader = csv.reader(f, dialect='semicolons')
    wb.create_sheet(shname)
    ws = wb.worksheets[shnum]
    #ws.title = "vrtyGIS"
    for row_index, row in enumerate(reader):
        for column_index, cell in enumerate(row):
            # column_letter = get_column_letter((column_index + 1))
            try:
                ws.cell(column=column_index+1, row=row_index+1).value = cell
            except Exception as err:
                print("Exception:", __file__, __name__, f"{err=}, {type(err)=}", row)
                logger.error("Exception: %s %s %s", err, type(err), row)


    f.close

create_xls_from_cvs(GEO5QGIS, SHGEO5,1)
create_xls_from_cvs(VRTCSVQGIS, SHVRT, 2)
create_xls_from_cvs(VRTLOZCSVQGIS, SHVRTLOZ, 3)


wb.save(filename = dest_filename)