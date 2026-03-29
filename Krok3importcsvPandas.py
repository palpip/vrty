import csv
import os
import logging
import pandas as pd
import openpyxl as opx
from _funcs import *
from _settings import *
chkdirs()



# logging.basicConfig(filename = KROK3LOGFILE, level=logging.INFO, filemode='w')
# logger=logging.getLogger()
logging.basicConfig(filename = KROK3LOGFILE, level=logging.DEBUG, filemode='a')
logger=logging.getLogger('Krok3')
logger.addHandler(logging.StreamHandler())

TOPDIR = TOPDIR + '\\'
pd.DataFrame().to_excel(EXCELWB, sheet_name = 'info', index=0)

def save_frame(df, dirname, dfname):
    '''ulozi dataframes v adresari dirname vo formate
    dfname.csv, dfname.xlsx, dfname.sqlite3
    excelovské súbory sa ukladajú do XLSNAME '''
    
    # print(f'ukladám {dfname}')
    if len(df) < 1.048e6:
        try:    
            with pd.ExcelWriter(EXCELWB, mode = 'a', engine='openpyxl', if_sheet_exists='replace') as excelwriter:
                df.to_excel(excelwriter, sheet_name=dfname)
                #logger.info(f"{dirname} - {dfname} {len(df)} riadkov uložených do excelu {XLSNAME}:{dfname}")
                logger.info(f"{len(df)} riadkov uložených do excelu {EXCELWB}:{dfname}")
                
                # df.info()
        except Exception as err:
                # print("Exception:", __file__, __name__, f"{err=}, {type(err)=}")
                logger.error("Exception: %s %s", err, type(err))
    else:
        logger.error(f'{dfname} obsahuje {len(df)} riadkov, viac ako je povolený limit excelu, nie je uložené')


# def create_xls_from_cvs(fn, shname, shnum):
#     print(fn)
#     f = open(fn)
#     csv.register_dialect('semicolons', delimiter=';')
#     reader = csv.reader(f, dialect='semicolons')
#     wb.create_sheet(shname)
#     ws = wb.worksheets[shnum]
#     #ws.title = "vrtyGIS"
#     for row_index, row in enumerate(reader):
#         for column_index, cell in enumerate(row):
#             # column_letter = get_column_letter((column_index + 1))
#             try:
#                 ws.cell(column=column_index+1, row=row_index+1).value = cell
#             except Exception as err:
#                 print("Exception:", __file__, __name__, f"{err=}, {type(err)=}", row)
#                 logger.error("Exception: %s %s %s", err, type(err), row)


#     f.close


def create_xls_from_cvsPandas(fn, shname, shnum):
    if os.path.isfile(fn):
        df = pd.read_csv(fn, on_bad_lines='warn', delimiter=';', decimal='.', encoding='ANSI')
        df = to_num(df, ['JTSKX', 'JTSKY'])
        df = clean_duplicates(df)
        bad_coor = bad_coordinates_df(df)
        df = good_coordinates_df(df)
    else:
        logger.warn(f'Adresár {TOPDIR} neobsahuje súbory {shname}')
    save_frame(bad_coor, TOPDIR, 'ERR' + shnum) #v kazdom pripade vytvor v exceli prazdne sheety
    save_frame(df, TOPDIR, shnum)

def create_xls_from_cvsPandas2(fn, shname, shnum):
    print(fn)
    df = bad_coor_df = pd.DataFrame() # dolezite pre vytvorenie prazdneho sheetu v xlsx v spodnom riadku
    if os.path.isfile(fn):
        try:
            df = pd.read_csv(fn, on_bad_lines='warn', delimiter=';', decimal='.', encoding='ANSI')
            if (read_count := df.shape[0]) != 0: 
                df = to_num(df, ['JTSKX', 'JTSKY'])
                df = clean_duplicates(df, shname)
                dedupl_count = df.shape[0]
                bad_coor_df = bad_coordinates_df(df)
                df = good_coordinates_df(df)
                good_coor_count = df.shape[0]
                # print(df.shape)
                # print(df.Vrt.head())
                # print(df)
                logger.info(f'súbor {fn}: načítaných {read_count} deduplikovaných {dedupl_count}' \
                            f' so správnymi súr. {good_coor_count} riadkov')
            else:
                logger.info(f'súbor {fn}: neobsahuje dáta')
        except Exception as err:
            logger.info(f'súbor {fn}: neobsahuje list {shname} {err}')
    else:
        logger.warn(f'Adresár {TOPDIR} neobsahuje súbor {shname}')
    save_frame(df, TOPDIR, shnum)
    save_frame(bad_coor_df, TOPDIR, 'ERR' + shnum)



    



logger.info(40*'+'+' started krok3 Pandas')
create_xls_from_cvsPandas2(VRTLOZCSVQGIS, SHVRTLOZ,SHVRTLOZ)
create_xls_from_cvsPandas2(VRTCSVQGIS, SHVRT, SHVRT)
create_xls_from_cvsPandas2(GEO5QGIS, SHGEO5, SHGEO5)
logger.info(40*'+'+' done krok3 Pandas')


